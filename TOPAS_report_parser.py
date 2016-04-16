#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
import re
import os
import logging
import time
from collections import Counter
import tkinter as tk
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox

logger = logging.getLogger('parser')
logger.setLevel(logging.ERROR)
fh = logging.FileHandler('TOPAS_report_parser.log')
fh.setLevel(logging.DEBUG)
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
logger.addHandler(fh)


def duplicates(lst, item):
    return [i for i, x in enumerate(lst) if x == item]


def try_float(s):
    try:
        r = float(s)
    except:
        r = s
    return r


class App():

    def __init__(self, master):
        self.master = master
        self.files = []
        self.files_path = ''
        self.output_path = ''
        helper_text = '''
1.Select files(txt only) you want to parse, you can choose multi-files;
2.Select the directory where excel file would be saved, duplicates will backup.
        '''
        self.helper = tk.Label(master=self.master, text=helper_text)
        self.helper.grid(row=0, columnspan=2)
        self.mes = tk.Label(master=self.master, text='')
        self.mes.grid(row=1, columnspan=2)
        self.files_getter = tk.Button(
            master=self.master, text="Inputs", command=self.get_files)
        self.files_getter.grid(row=2, column=0)
        self.file_putter = tk.Button(
            master=self.master, text="Output", command=self.get_output_path)
        self.file_putter.grid(row=2, column=1)
        self.parser_pattern = re.compile(
            r'[\s\S]*R-Values [\s\S]*Rwp : (\d+\.\d+)[\s\S]*?Rp[\s\S]*Quantitative Analysis - Rietveld ([\s\S]*)Background[\s\S]*', re.M)
        self.parser_phase_pattern = re.compile(
            r' *Phase \d.*: (.*?) *?(\d+\.\d+) %.*')

    def get_files(self):
        self.mes.config(text='')
        self.mes.bind(sequence="<Button-1>", func=None)
        files = [os.path.normpath(file)
                 for file in filedialog.askopenfilenames(defaultextension='.txt', filetypes=[('text files', '.txt')])]
        self.files = files
        try:
            self.files_path = os.path.split(files[0])[0]
        except:
            pass
        logger.info('Selected files:{:s}'.format('\n'.join(self.files) + '\n'))

    def get_output_path(self):
        opts = {'defaultextension': '.xlsx', 'filetypes': [
            ('Excel work book', '.xlsx')], 'initialdir': self.files_path}
        path = filedialog.asksaveasfilename(**opts)
        self.output_path = os.path.normcase(path)
        logger.info('Selected a save path:{:s}'.format(self.output_path))
        if not self.files:
            return
        if self.output_path:
            self.output_result()

    def output_result(self):
        try:
            logger.info('Saving to path:{:s}\nWith these files:{:s}'.format(
                self.output_path, '\n'.join(self.files) + '\n'))
            if os.path.exists(self.output_path):
                os.rename(self.output_path, os.path.splitext(self.output_path)[
                          0] + '_backup_' + str(time.time()) + '.xlsx')
            result = [self.parser(file) for file in self.files]
            headers = [h for (h, _) in sum((Counter(i.keys()) for i in result), Counter(
            )).most_common()]  # make sure the item sort by counts
            headers.remove('ID')
            headers.remove('Rwp')
            headers.insert(0, 'Rwp')
            headers.insert(0, 'ID')
            columns = {k: i for i, k in enumerate(headers, start=1)}
            work_book = openpyxl.Workbook()
            logger.info('Opened a work book')
            work_sheet = work_book.active
            #write headers
            for header, column in columns.items():
                work_sheet.cell(row=1, column=column, value=header)
            for row, r in enumerate(result, start=2):
                logger.info('resulting')
                for k, value in r.items():
                    work_sheet.cell(row=row, column=columns[
                                    k], value=try_float(value))
            work_book.save(self.output_path)
            logger.info('Saved to path:{:s}\nWith these files:{:s}'.format(
                self.output_path, '\n'.join(self.files) + '\n'))
            self.mes.config(
                text='Done! -->Click<-- to open the dir.', foreground='red')
            self.mes.bind(
                sequence="<Button-1>", func=lambda e: os.system('explorer ' + os.path.split(self.output_path)[0]))
        except Exception as e:
            messagebox.askquestion(
                title='Error!', message='Can not save as this file name, pick another name pls!')
            logger.info('Faild to save as:{:s}'.format(self.output_path) +
                        'Because:' + str(e))

    def parser(self, file):
        with open(file, 'r') as f:
            t = self.parser_pattern.match(f.read())
        file_id = os.path.splitext(os.path.split(file)[1])[0]
        #deal with duplicat phase names
        phase_list = [[item.strip('"_ \'\\\/?\t.') for item in i.groups()] for i in filter(
            None, (self.parser_phase_pattern.match(line) for line in t.groups()[1].splitlines()))]
        phase_name = tuple(phase[0] for phase in phase_list)
        for dt in set(phase_name):
            dups = duplicates(phase_name, dt)
            if len(dups) > 1:
                for j, k in enumerate(dups):
                    phase_list[k][0] += '_' + str(j)
        temp = {'ID': file_id, 'Rwp': t.groups()[0]}
        temp.update(dict(phase_list))
        return temp

if __name__ == "__main__":
    title = ' TOPAS report parser by Lilin Lao'
    root = tk.Tk(className=title)
    root.wm_title = title
    root.title = title
    app = App(master=root)
    root.mainloop()

